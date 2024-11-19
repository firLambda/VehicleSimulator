import os
import csv
import generate
from tqdm import tqdm
import datetime
import openpyxl as px
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font

# 任意の車線を走行した車両のIDが入ったリストを返す関数


def search_carid_lane(car, car_max, time_max, lane):
  id_list = []
  for i in range(car_max):
    for j in range(time_max):
      if car[j][i].lane == lane:
        id_list.append(i)
        break
  return id_list


def save4_write0(ws, car, time_max, lane, lane_id_list, interval, acceleration_lane_start, acceleration_lane_end):  # save4専用の関数
  sum1 = 0
  sum2 = 0
  t1 = 0
  t2 = 0
  ws.freeze_panes = 'A2'  # 先頭行固定
  if 0 in lane_id_list:  # id=0が含まれていた場合
    lane_id_list.remove(0)
  row_tmp = 1
  for j, k in enumerate(lane_id_list, 1):
    ws.cell(row=row_tmp, column=j).value = (k)  # 一番上の行にIDを記録する
  row_tmp += 1
  for i in range(time_max):
    if i % interval == 0:
      for j, k in enumerate(lane_id_list, 1):
        if car[i][k].lane == lane:
          ws.cell(row=row_tmp, column=j).value = (car[i][k].vel * 3.6)
          if acceleration_lane_start - 300 < car[i][k].front < acceleration_lane_end:
            sum1 += car[i][k].vel * 3.6
            t1 += 1
          if acceleration_lane_start < car[i][k].front < acceleration_lane_end:
            sum2 += car[i][k].vel * 3.6
            t2 += 1
        else:
          ws.cell(row=row_tmp, column=j).value = ("")
      row_tmp = row_tmp + 1
  ws.cell(row=row_tmp, column=1).value = ("平均速度")
  ws.cell(row=row_tmp, column=2).value = (sum1 / t1)
  ws.cell(row=row_tmp, column=3).value = (sum2 / t2)


def abc_from_number(number):  # 可視化するのに使う関数
  alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
  colname = ""
  divend = number
  while divend > 0:
    modulo = (divend - 1) % 26
    colname = alphabet[modulo] + colname
    divend = (divend - modulo) // 26
  return colname


def colorBarRGB(car_id):  # 可視化するのに使う関数
  car_id = car_id % 10
  if car_id == 1:
    color = "FF0000"  # 赤
  elif car_id == 2:
    color = "FFA500"  # オレンジ
  elif car_id == 3:
    color = "00FF00"  # 黄緑
  elif car_id == 4:
    color = "007400"  # 緑
  elif car_id == 5:
    color = "00FFFF"  # 水色
  elif car_id == 6:
    color = "0000FF"  # 青
  elif car_id == 7:
    color = "8D0093"  # 紫
  elif car_id == 8:
    color = "FF00FF"  # ピンク
  elif car_id == 9:
    color = "800000"  # 茶色
  elif car_id == 0:
    color = "808080"  # グレー
  return (color)


def create_path(seed, q_lane1, q_lane2, use_base_station, ratio):
  now = datetime.datetime.now()  # 現在時刻取得
  date = str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(now.minute)  # 201910301234みたいに作る
  if use_base_station == 0:
    path = os.getcwd() + "/" + "seed" + "_" + str(seed) + "車両の割合" + str(ratio) + "%q1_" + str(round(q_lane1, 2)) + "_q2_" + str(round(q_lane2, 2)) + "基地局なし" + "_" + str(date) + ".xlsx"
  else:
    path = os.getcwd() + "/" + "seed" + "_" + str(seed) + "車両の割合" + str(ratio) + "%q1_" + str(round(q_lane1, 2)) + "_q2_" + str(round(q_lane2, 2)) + "基地局あり" + "_" + str(date) + ".xlsx"
  return path


def create_excel_file():
  wb = px.Workbook()
  ws = wb.active
  return wb, ws


def create_info_sheet(ws, time_max, q_lane1, q_lane2, q_lane3, acceleration_lane_start, acceleration_lane_end, use_base_station, ratio):
  ws.cell(row=1, column=1).value = ("シミュレーション時間")
  ws.cell(row=1, column=2).value = (time_max / 10)
  ws.cell(row=2, column=1).value = ("加速交通量")
  ws.cell(row=2, column=2).value = (q_lane1)
  ws.cell(row=3, column=1).value = ("走行交通量")
  ws.cell(row=3, column=2).value = (q_lane2)
  ws.cell(row=4, column=1).value = ("追越交通量")
  ws.cell(row=4, column=2).value = (q_lane3)
  ws.cell(row=5, column=1).value = ("加速車線開始[m]")
  ws.cell(row=5, column=2).value = (acceleration_lane_start)
  ws.cell(row=6, column=1).value = ("加速車線終了[m]")
  ws.cell(row=6, column=2).value = (acceleration_lane_end)
  ws.cell(row=7, column=1).value = ("基地局")
  ws.cell(row=7, column=2).value = (use_base_station)
  ws.cell(row=8, column=1).value = ("譲る車両の割合")
  ws.cell(row=8, column=2).value = (ratio)
  return ws


def create_lane_vel_sheet(wb, ws, car, car_max, time_max, interval, acceleration_lane_start, acceleration_lane_end):
  lane1_id_list, lane2_id_list, lane3_id_list = [], [], []
  lane1_id_list = search_carid_lane(car, car_max, time_max, 1)  # 加速車線を走行した車両のIDを調べる
  lane2_id_list = search_carid_lane(car, car_max, time_max, 2)  # 走行車線を走行した車両のIDを調べる
  lane3_id_list = search_carid_lane(car, car_max, time_max, 3)  # 追越車線を走行した車両のIDを調べる
  ws = wb.create_sheet(title="加速速度")  # シート名を指定
  save4_write0(ws, car, time_max, 1, lane1_id_list, interval, acceleration_lane_start, acceleration_lane_end)
  ws = wb.create_sheet(title="走行速度")  # シート名を指定
  save4_write0(ws, car, time_max, 2, lane2_id_list, interval, acceleration_lane_start, acceleration_lane_end)
  ws = wb.create_sheet(title="追越速度")  # シート名を指定
  save4_write0(ws, car, time_max, 3, lane3_id_list, interval, acceleration_lane_start, acceleration_lane_end)
  return ws


def creaet_merging_info_sheet(wb, ws, car, car_max, time_max, base_station, interval, car_info, acceleration_lane_start, acceleration_lane_end):
  # ここから加速車線から走行車線に移動するまでにかかった時間を記録する
  ws = wb.create_sheet(title="合流にかかった時間")
  fill = PatternFill(patternType='solid', fgColor='FF0000')
  row_tmp = 2
  ws.cell(row=row_tmp, column=1).value = ("ID")
  ws.cell(row=row_tmp, column=2).value = ("車線変更時刻")
  ws.cell(row=row_tmp, column=3).value = ("車両生成から車線変更終了までの時間")
  ws.cell(row=row_tmp, column=4).value = ("位置")
  ws.cell(row=row_tmp, column=5).value = ("譲った車両")
  ws.cell(row=row_tmp, column=6).value = ("検索前方位置")
  ws.cell(row=row_tmp, column=7).value = ("検索後方位置")
  ws.cell(row=row_tmp, column=8).value = ("合流時車両速度")
  ws.cell(row=row_tmp, column=9).value = ("1秒前合流時車両速度")
  ws.cell(row=row_tmp, column=10).value = ("2秒前合流時車両速度")
  ws.cell(row=row_tmp, column=11).value = ("3秒前合流時車両速度")
  ws.cell(row=row_tmp, column=12).value = ("4秒前合流時車両速度")
  ws.cell(row=row_tmp, column=13).value = ("希望速度")
  ws.cell(row=row_tmp, column=14).value = ("車両発生時刻")
  row_tmp += 1
  time_tmp = 0
  sum1 = 0
  sum2 = 0
  sum_tmp = 0
  # ##合流できた車両の合流までにかかった時間を調べる###
  for i in range(1, car_max):  # 車両IDが0の車は除く
    for j in range(time_max):
      if j % interval == 0:
        if acceleration_lane_start < car[j][i].back:
          if car[j][i].lane == 1:
            time_tmp += 1  # 加速車線を走った時間を記録している
          elif car[j][i].lane == 2:
            break  # 走行車線を走った場合はそれ以上調べる必要はない
          elif car[j][i].lane == 3:
            break  # 追越車線を走った場合はそれ以上調べる必要はない
          else:
            break
    # 加速車線を走っていた時間が0秒より長く、最後の時間の車線が加速車線でないとき、合流できた車両とみなす
    if time_tmp > 0 and car[j][i].lane != 1:
      ws.cell(row=row_tmp, column=1).value = (car_info[i].id)  # 合流できた車両ID
      ws.cell(row=row_tmp, column=2).value = (j / 10)  # 合流した時刻
      ws.cell(row=row_tmp, column=3).value = (time_tmp / 10)  # 合流できるまでにかかった時間
      ws.cell(row=row_tmp, column=4).value = (car[j][i].front)  # 合流した位置
      if acceleration_lane_end - 50 < car[j][i].front:  # 950m以降の車線変更ではセル色を変える
        ws.cell(row=row_tmp, column=4).fill = fill
      ws.cell(row=row_tmp, column=5).value = (base_station[int(car[j][i].id)][2])  # 譲った車両
      ws.cell(row=row_tmp, column=6).value = (base_station[int(car[j][i].id)][0])  # 検索前方位置
      ws.cell(row=row_tmp, column=7).value = (base_station[int(car[j][i].id)][1])  # 検索後方位置
      for k in range(0, 5):
        ws.cell(row=row_tmp, column=(8 + k)).value = (car[j - (k * 10)][i].vel * 3.6)  # 4秒前までの合流時車両速度
        if car[j][i].vel < 20 / 3.6 and k == 0:  # 20km/s以下の車線変更ではセル色を変える
          ws.cell(row=row_tmp, column=8).fill = fill
      ws.cell(row=row_tmp, column=13).value = (car_info[i].vd * 3.6)  # 希望速度
      ws.cell(row=row_tmp, column=14).value = (car_info[i].occur_time / 10)  # 車両発生時刻
      if j >= 3000:
        sum1 += time_tmp
        sum2 += car[j][i].front
        sum_tmp += 1
      row_tmp += 1
    time_tmp = 0
  if(sum_tmp != 0):
    ave_time = sum1 / sum_tmp
    ave_length = sum2 / sum_tmp
    ws.cell(row=row_tmp + 1, column=1).value = ("300秒以降の車両の時間平均値")
    ws.cell(row=row_tmp + 1, column=2).value = (ave_time / 10)
    ws.cell(row=1, column=1).value = (ave_time / 10)
    ws.cell(row=row_tmp + 2, column=1).value = ("300秒以降の車両の距離平均値")
    ws.cell(row=row_tmp + 2, column=2).value = (ave_length - acceleration_lane_start)
    ws.cell(row=1, column=2).value = (ave_length - acceleration_lane_start)
  row_tmp = 3
  time_tmp = 0
  sum1 = 0
  sum2 = 0
  sum_tmp = 0
  for i in range(1, car_max):  # 車両IDが0の車は除く
    for j in range(time_max):
      if j % interval == 0:
        if car[j][i].lane == 1:
          time_tmp += 1  # 加速車線を走った時間を記録している
        elif car[j][i].lane == 2:
          break  # 走行車線を走った場合はそれ以上調べる必要はない
        elif car[j][i].lane == 3:
          break  # 追越車線を走った場合はそれ以上調べる必要はない
    # 加速車線を走っていた時間が0秒より長く、最後の時間の車線が加速車線でないとき、合流できた車両とみなす
    if time_tmp > 0 and car[j][i].lane != 1:
      if j >= 3000:
        sum1 += (time_tmp - ave_time)**2
        sum2 += (car[j][i].front - ave_length)**2
        sum_tmp += 1
      row_tmp += 1
    time_tmp = 0
  if(sum_tmp != 0):
    ws.cell(row=row_tmp + 3, column=1).value = ("300秒以降の車両の時間平均値の標準偏差")
    ws.cell(row=row_tmp + 3, column=2).value = ((sum1 / sum_tmp)**0.5)
    ws.cell(row=row_tmp + 4, column=1).value = ("300秒以降の車両の距離平均値の標準偏差")
    ws.cell(row=row_tmp + 4, column=2).value = ((sum2 / sum_tmp)**0.5)
  return ws


def create_visual_sheet(wb, ws, car, car_max, time_max, interval, acceleration_lane_start, acceleration_lane_end, road_length):
  border = Border(right=Side(style='thin', color='000000'))
  ws = wb.create_sheet(title="可視化")
  lane_id_list = [0] * 2000
  fill = PatternFill(patternType='solid', fgColor='FF0000')
  for i in tqdm(range(time_max), desc='save_cal'):
    if i % interval == 0:
      ws.cell(row=1, column=1 + 3 * i / interval).value = (i / interval)  # IDを記録1
      for j in range(1, car_max):
        if 3000 > car[i][j].front > 0 and car[i][j].lane > 0:
          lane = car[i][j].lane
          color_tmp = colorBarRGB(car[i][j].id)
          ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).font = Font(b=True, color=color_tmp)
          if car[i][j].lane == 3:
            ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).value = (str(j) + str("/") + str(int(car[i][j].vel * 3.6)))  # IDを記録1
            ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).border = border
          else:
            ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).value = (str(j) + str("/") + str(int(car[i][j].vel * 3.6)))  # IDを記録1
          # 車線が変わった時だけセル色を変える
          if lane != lane_id_list[j] and lane_id_list[j] != 0:
            ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).fill = fill
          if car[i][j].shift_begin + 30 < i and lane == 1 and car[i][j].shift_lane == 1:
            ws.cell(row=car[i][j].front, column=car[i][j].lane + 3 * i / interval).fill = fill
          lane_id_list[j] = lane
      for j in range(1, road_length):  # 罫線引くところ 20200615
        ws.cell(row=j, column=3 + 3 * i / interval).border = border
  for i in range(time_max):
    col = abc_from_number(i + 1)
    ws.column_dimensions[col].width = 8
  ws.freeze_panes = 'A2'  # 先頭行固定
  return ws


def create_log_sheet(wb, ws, car, car_max, time_max, interval, car_info):
  ws = wb.create_sheet(title="log")
  row_tmp = 1
  # 0ID、1前方位置、2後方位置、3車線、4速度、5加速度、6車間距離,7相対速度,8前方車両ID
  # 9車線変更している途中か,10車線変更先の車線,11車線変更開始時間,12車線変更先車間距離,13車線変更先前方車両ID,14目標車両ID,15希望速度
  list = ["時間", "ID", "前方位置", "後方位置", "車線", "速度", "加速度", "車間",
          "相対速度", "前方車両ID", "車線変更途中", "車線変更先", "車線変更開始時間",
          "車線変更先車間距離", "車線変更先前方車両ID", "目標車両ID", "希望速度"]
  for col_tmp, list_tmp in enumerate(list, 1):  # 1行目の文字の部分の書き込み col_tmpは1からスタートする
    ws.cell(row=row_tmp, column=col_tmp).value = (list_tmp)
  row_tmp = 2
  ws.freeze_panes = 'A2'  # 先頭行固定
  for time_tmp in range(0, time_max, int(interval)):
    for car_tmp in range(1, car_max):  # ID0は除外,車の台数分ループ
      if car[time_tmp][car_tmp].lane == 0:
        continue
      list = car[time_tmp][car_tmp].vars_to_list()  # listは1次元
      ws.cell(row=row_tmp, column=1).value = (time_tmp / 10)  # 最も左の列の時間の部分
      for col_tmp, list_tmp in enumerate(list, 2):  # change by kino 2020-01-28
        if col_tmp == 6 or col_tmp == 9:  # 速度,相対速度
          ws.cell(row=row_tmp, column=col_tmp).value = (list_tmp * 3.6)  # 速度のときはm/sからkm/hに変更
        else:
          if col_tmp == 16:
            ws.cell(row=row_tmp, column=17).value = (car_info[car_tmp].vd * 3.6)
          ws.cell(row=row_tmp, column=col_tmp).value = (list_tmp)
      row_tmp += 1
  return ws

# excel直接出力（加速・走行・追い越し用）


def save4(car, base_station, use_base_station, car_max, time_max, interval, q_lane1, q_lane2, q_lane3, acceleration_lane_start, acceleration_lane_end, seed, car_info, ratio, road_length):
  path = create_path(seed, q_lane1, q_lane2, use_base_station, ratio)
  wb, ws = create_excel_file()
  ws.title = "情報"  # 最初は'sheet1'というシートがデフォルトで作成されるため、名前変更
  ws = create_info_sheet(ws, time_max, q_lane1, q_lane2, q_lane3, acceleration_lane_start, acceleration_lane_end, use_base_station, ratio)
  ws = create_lane_vel_sheet(wb, ws, car, car_max, time_max, interval, acceleration_lane_start, acceleration_lane_end)
  ws = creaet_merging_info_sheet(wb, ws, car, car_max, time_max, base_station, interval / 10, car_info, acceleration_lane_start, acceleration_lane_end)
  ws = create_visual_sheet(wb, ws, car, car_max, time_max, interval, acceleration_lane_start, acceleration_lane_end, road_length)  # 可視化
  ws = create_log_sheet(wb, ws, car, car_max, time_max, interval, car_info)  # logデータ出力
  print(path + "を保存中")
  wb.save(path)
